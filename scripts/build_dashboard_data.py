"""Build dashboard data JSON for the 2026 KEQAS EQA Dashboard."""
import json
import openpyxl
from collections import Counter

# Load main data
with open('data/dashboard_data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Load labs data
wb = openpyxl.load_workbook('data/2026 KEQAS EQA labs.xlsx', read_only=True)
ws = wb[wb.sheetnames[0]]

cat_en = {
    '수혈의학': 'Transfusion Medicine',
    '진단혈액학': 'Diagnostic Hematology',
    '임상화학': 'Clinical Chemistry',
    '진단면역학': 'Diagnostic Immunology',
    '임상미생물학': 'Clinical Microbiology',
    '진단유전학': 'Diagnostic Genetics',
    '검사실운영': 'Laboratory Operations'
}

# Build programs list
programs_list = []
for cat, subcats in data['categories_tree'].items():
    for subcat, progs in subcats.items():
        for p in progs:
            prog = {
                'category': cat_en.get(cat, cat),
                'category_kr': cat,
                'subcategory': p.get('classification', subcat),
                'subcategory_kr': subcat,
                'name': p.get('name', ''),
                'name_en': p.get('program_en', ''),
                'sample_type': p.get('sample_type', ''),
                'level': p.get('level', ''),
                'dist_per_yr': p.get('dist_per_yr', ''),
                'shipping': p.get('shipping', ''),
                'lypo_liquid': p.get('lypo_liquid', ''),
                'commutable': p.get('commutable', ''),
                'sourcing': p.get('sourcing', ''),
                'origin': p.get('origin', ''),
                'num_labs': p.get('num_labs', 0),
                'num_tests': len(p.get('tests', [])),
                'tests': p.get('tests', []),
                'guidebook': p.get('guidebook', [])
            }
            programs_list.append(prog)

# Build labs classification data
detail_en = {
    '상급종합병원': 'Advanced General Hospital',
    '종합병원': 'General Hospital (100+)',
    '일반병원(치과, 한방, 요양병원)': 'General Hospital',
    '전문병원': 'Specialty Hospital',
    '일반의원(치과, 한의원)': 'General Clinic',
    '보건소': 'Public Health Center',
    '군병원(의무근무대)': 'Military Hospital',
    '전문수탁기관': 'Specialized Reference Lab',
    '공익 목적의 비의료기관': 'Non-Medical (Public Interest)',
    '비의료기관(연구소 등)': 'Non-Medical (Research)'
}

new_class_en = {
    '의원(30병상 이하)': 'Primary',
    '병원(30~100병상)': 'Secondary (Small)',
    '종합병원(100병상 이상)': 'Secondary (Large)',
    '군병원(의무근무대)': 'Military',
    '보건소': 'Public',
    '공익 목적의 비의료기관': 'Non-Medical',
    '비의료기관(연구소 등)': 'Non-Medical'
}

# Count classifications
detail_counter = Counter()
new_class_counter = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    detail_counter[row[4]] += 1
    new_class_counter[row[3]] += 1

labs_data = []
# Group by new classification (broad), with details
broad_groups = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    broad = row[3]
    detail = row[4]
    if broad not in broad_groups:
        broad_groups[broad] = Counter()
    broad_groups[broad][detail] += 1

labs_treemap = []
for broad, details in broad_groups.items():
    broad_en = new_class_en.get(broad, broad)
    children = []
    for detail, count in details.most_common():
        children.append({
            'name': detail_en.get(detail, detail),
            'name_kr': detail,
            'count': count,
            'pct': round(count / 2119 * 100, 1)
        })
    labs_treemap.append({
        'name': broad_en,
        'name_kr': broad,
        'count': sum(details.values()),
        'pct': round(sum(details.values()) / 2119 * 100, 1),
        'children': children
    })

labs_treemap.sort(key=lambda x: x['count'], reverse=True)

# Build output
output = {
    'programs': programs_list,
    'labs': {
        'total': 2119,
        'treemap': labs_treemap,
        'detail_counts': [{
            'name': detail_en.get(k, k),
            'name_kr': k,
            'count': v,
            'pct': round(v / 2119 * 100, 1)
        } for k, v in detail_counter.most_common()]
    },
    'summary': {
        'total_programs': 103,
        'total_tests': 548,
        'total_labs': 2119,
        'num_categories': 7
    }
}

with open('data/dashboard_embed.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"Generated dashboard_embed.json: {len(programs_list)} programs, {len(labs_treemap)} lab groups")
